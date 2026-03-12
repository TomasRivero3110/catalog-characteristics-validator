# -*- coding: latin-1 -*-
#! / usr / bin / python
# vim: set fileencoding = latin-1:
import os, sys
import subprocess
import unicodedata
import tkinter.filedialog  #Libreria para crear cuadros de dialogo
import csv, sys, re, datetime, os #Librerias de python y del sistema
import xlwt, xlrd     #Librerias de gestion de archivos excel
import sys
import openpyxl
import pandas as pd
import datetime
from openpyxl.utils import get_column_letter
import re
from dateutil import parser
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'ScriptValidacionCaracteristicas')))
from consultaBDDEquiposOnline import Online   #impartando funciones de modulo de BD
from consultaBDDEquiposForPartners import EquiposFP
from consultaBDDEquipos import Equipos
from consultaLoyalty import Loyalty
from openpyxl.workbook import Workbook as openpyxlWorkbook
from openpyxl.styles import PatternFill, Alignment, Font

lista_id_equipos_TEST = [] #lista de equipos de la base
lista_id_equiposOnline_TEST = []
lista_id_equiposFP_TEST = []
lista_planilla = [] #planilla de MKT con caracteristicas especificadas 
lista_rangos = []
lista_loyalty = []
tipo = ""
#estilo del Excel de salida
workbook2 = xlwt.Workbook('ascii',style_compression=2) 
xlwt.add_palette_colour("purpura_custom", 0x21)
workbook2.set_colour_RGB(0x21, 204, 204, 255)
xlwt.add_palette_colour("verde_custom", 0x22)
workbook2.set_colour_RGB(0x22, 0, 255, 0)
xlwt.add_palette_colour("red_custom", 0x23)
workbook2.set_colour_RGB(0x23, 255, 102, 0)
xlwt.add_palette_colour("azul_custom", 0x24)
workbook2.set_colour_RGB(0x24, 0, 204, 255)
xlwt.add_palette_colour("amarillo_custom", 0x25)
workbook2.set_colour_RGB(0x25, 255, 255, 153)
sheet2 = workbook2.add_sheet("POI Worksheet",  cell_overwrite_ok=True)

def buscarPosiciones(titulo):
    global lista_planilla
    for i in range(len(lista_planilla[2])):
        if str(titulo).lower() == str(lista_planilla[1][i]).lower():
            #print(titulo,i)
            return i   
    mensaje = "No se encontro el titulo " + titulo + " en la BDD"
    sys.exit(mensaje)

def openFileDialog():
    """Esta funcion abre un cuadro de dialogo para cargar archivos."""
    root = tkinter.Tk()
    root.withdraw()
    file_path = tkinter.filedialog.askopenfilename()
    if not file_path:  # Verificar si el usuario no seleccionó ningún archivo
        print("No se selecciono ningun archivo.")
        exit()  # Terminar el programa si no se seleccionó ningún archivo
    return file_path
    
def saveFileDialog():
    '''Esta funcion abre un cuadro de dialogo para guardar archivos''' 
    ftypes = [ ('Libro de Excel','*.xlsx'), ('Libro de Excel 97-2003','*.xls'), ('All files', '*'),  ]
    root = tkinter.Tk()
    root.withdraw()
    file_path = tkinter.filedialog.asksaveasfilename(filetypes=ftypes)
    return file_path + ".xls"

def readCSVPlanes(file_path):
    # Crear una lista vacía para almacenar las filas como listas
    lista_filas = []
    # Cargar el libro de trabajo de Excel
    wb = openpyxl.load_workbook(file_path, data_only=True)
    # Obtener la hoja activa
    sheet = wb.active
    # Iterar sobre las filas de la hoja
    for row in sheet.iter_rows(values_only=True):
        # Agregar la fila como lista a la lista de filas
        lista_filas.append(list(row))
    # Cerrar el libro de trabajo para liberar recursos
    wb.close()
    # Devolver el DataFrame y la lista de filas
    return lista_filas

def guardarComoXLSX(filepath):
    '''Guardar el archivo como .XLSX, borrar el .XLS generado y aplicar estilos/anchos'''
    xlsBook = xlrd.open_workbook(filepath)
    workbook2 = openpyxlWorkbook()
    sheet2 = None
    xrange = range

    for i in xrange(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet2 = workbook2.active if i == 0 else workbook2.create_sheet()
        sheet2.title = xlsSheet.name

        # Copiar valores
        for row in xrange(0, xlsSheet.nrows):
            for col in xrange(0, xlsSheet.ncols):
                cell = sheet2.cell(row=row + 1, column=col + 1)
                cell.value = xlsSheet.cell_value(row, col)

        # Estilos de encabezado (fila 1)
        max_row = sheet2.max_row
        max_col = sheet2.max_column
        for c in range(1, max_col + 1):
            h = sheet2.cell(row=1, column=c)
            h.font = Font(name='Calibri', size=11, color="000000", bold=True)
            h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Coloreo condicional de "Result" (columna J = 10) y wrap en "Error" (columna K = 11)
        from openpyxl.styles import PatternFill
        FILL_GREEN = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")  # verde claro
        FILL_RED   = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")  # rojo claro

        for r in range(2, max_row + 1):
            result_cell = sheet2.cell(row=r, column=6)  # "Result"
            error_cell  = sheet2.cell(row=r, column=7)  # "Error" (descripción)
            # Asegurar wrap en la descripción de error
            error_cell.alignment = Alignment(wrap_text=True, vertical="top")

            val = str(result_cell.value).strip().lower() if result_cell.value is not None else ""
            if val == "ok":
                result_cell.fill = FILL_GREEN
            elif val == "error":
                result_cell.fill = FILL_RED

        # Ajuste automático de ancho de columnas (con límite)
        for c in range(1, max_col + 1):
            max_len = 0
            for r in range(1, max_row + 1):
                v = sheet2.cell(row=r, column=c).value
                if v is None:
                    continue
                l = len(str(v))
                if l > max_len:
                    max_len = l
            # ancho aproximado: largo + margen; límite para no exagerar
            sheet2.column_dimensions[get_column_letter(c)].width = min(60, max_len + 2)

        # Congelar encabezado
        sheet2.freeze_panes = "A2"

    filepath2 = filepath[:-4] + '.xlsx'
    workbook2.save(filepath2)

    if os.path.exists(filepath):
        os.remove(filepath)
    else:
        print("The file does not exist")

def ImprimirEncabezado():
    nombre_columnas = ['Offering Name' , 'Planilla Nombre' , 'Revenue code', 'Color', 'Colores', 'Result', 'Error'] 
    Font = "font: name Calibri, color-index black, height 220"
    for i in range(len(nombre_columnas)):
        sheet2.write(0, i, nombre_columnas[i], xlwt.easyxf("pattern: pattern solid, fore_color azul_custom; align: horiz center; "+ Font))
    return 0

def validar_clearing(cadena):
    if not cadena:  # None o vacío
        return False
    
    # Si ya es lista, usarla directamente
    if isinstance(cadena, list):
        partes = [str(p).strip() for p in cadena]
    else:
        partes = [p.strip() for p in str(cadena).split(",")]
    
    clearing_ok = False
    otras_ok = True

    for p in partes:
        if "Clearing" in p:
            clearing_ok = p.strip().endswith("(Si)")
        else:
            if not p.strip().endswith("(None)"):
                otras_ok = False
    
    return clearing_ok and otras_ok

def separarColores(value):
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        items = list(value)
    else:
        items = re.split(r"[,\;/\|]+", str(value))  # admite , ; / |
    out = []
    for x in items:
        if x is None:
            continue
        sx = str(x).strip()     # <- convierte ints a str antes de strip
        if sx:
            out.append(sx.upper())
    return out

def ImprimirBase(Offering, PlanillaNombre , MKT_Revenue_code, Color, Colores, j):
    sheet2.write(j, 0, str(Offering)) 
    sheet2.write(j, 1, str(PlanillaNombre)) 
    sheet2.write(j, 2, str(MKT_Revenue_code)) 
    sheet2.write(j, 3, str(Color)) 
    sheet2.write(j, 4, str(Colores))                                 
    
    
# def ImprimirBaseForPartners(Offering, Revenue_code , MKT_Revenue_code,Color, Colores, j):
#     sheet2.write(j, 0, str(Offering)) 
#     sheet2.write(j, 1, str(Revenue_code)) 
#     sheet2.write(j, 2, str(MKT_Revenue_code)) 
#     sheet2.write(j, 3, str(Color))                               
#     sheet2.write(j, 4, str(Colores))
    
def ValidarDatosForPartners(PlanillaNombre, PlanillaRC,PlanillaColores,PlanillaColorPrinicipal,equipo):
    respuesta = ['OK', '']
    NombreSistema = equipo[1]
    RevenueCodeSis = equipo[5]    
    ColoresSis = equipo[9]
    ColorSis=equipo[8]
    BulkTree = equipo[11]
    if(str(NombreSistema).lower() != PlanillaNombre):
        respuesta[0] = "Error"
        respuesta[1] = "Distinto nombre"
    if(str(RevenueCodeSis).lower() != PlanillaRC):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", distinto RC"
        else:
            respuesta[1] = "Revenue Code distinto"                
    if(str(ColorSis).lower() != PlanillaColorPrinicipal):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", distinto Default Color" 
        else:
            respuesta[1] = "Color principal distinto"                           
    if(str(BulkTree) != "None"):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Product Offering Associations erroneo" 
        else:
            respuesta[1] = "Product Offering Associations erroneo"                             
    if set(separarColores(ColoresSis)) != set(separarColores(PlanillaColores)):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", falla colores"
        else:
            respuesta[1] = "Colores definidos distintos"                
    return respuesta

def buscarEquipo(equipo_nombre):
       # Normalizar entradas a minúsculas
    equipo_nombre = str(equipo_nombre).lower() 
    for i in range(len(lista_id_equipos_TEST)): 
        # Normalizar valores de lista_id_equipos_TEST a minúsculas para la comparación
        nombre = str(lista_id_equipos_TEST[i][1]).lower()
        # Comparar con normalización y verificar si está en desarrollo
        if (equipo_nombre == nombre):
            return lista_id_equipos_TEST[i]
    
    # Formatear mensaje de error como una cadena
    #mensaje = 'error - No se encontro el Paquete'
    return None

def buscarLoyalty(equipo_nombre):
       # Normalizar entradas a minúsculas
    equipo_nombre = str(equipo_nombre).lower() 
    for i in range(len(lista_loyalty)): 
        # Normalizar valores de lista_id_equipos_TEST a minúsculas para la comparación
        nombre = str(lista_loyalty[i][1]).lower()
        # Comparar con normalización y verificar si está en desarrollo
        if (equipo_nombre == nombre):
            return lista_loyalty[i]
    
    # Formatear mensaje de error como una cadena
    #mensaje = 'error - No se encontro el Paquete'
    return None            

def buscarEquipoOnline(equipo_nombre):
       # Normalizar entradas a minúsculas
    equipo_nombre = str(equipo_nombre).lower()       
    for i in range(len(lista_id_equiposOnline_TEST)): 
        # Normalizar valores de lista_id_equipos_TEST a minúsculas para la comparación
        nombre = str(lista_id_equiposOnline_TEST[i][1]).lower()
        # Comparar con normalización y verificar si está en desarrollo
        if (equipo_nombre == nombre):
            return lista_id_equiposOnline_TEST[i]
    
    # Formatear mensaje de error como una cadena
    #mensaje = 'error - No se encontro el Paquete'
    return None
            

def buscarEquipoFP(equipo_nombre):
       # Normalizar entradas a minúsculas
    equipo_nombre = str(equipo_nombre).lower()       
    for i in range(len(lista_id_equiposFP_TEST)): 
        # Normalizar valores de lista_id_equipos_TEST a minúsculas para la comparación
        nombre = str(lista_id_equiposFP_TEST[i][1]).lower()
        # Comparar con normalización y verificar si está en desarrollo
        if (equipo_nombre == nombre):
            return lista_id_equiposFP_TEST[i]
    
    # Formatear mensaje de error como una cadena
    #mensaje = 'error - No se encontro el Paquete'
    return None
            
            
def _split_comas(value):
    # Acepta str o lista/tupla; devuelve lista limpia de tokens
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        items = list(value)
    else:
        items = str(value).split(",")
    out = []
    for x in items:
        sx = str(x).strip()
        if sx:
            out.append(sx)
    return out

def validar_nrc_unico(texto):
    partes = _split_comas(texto)
    return len(partes) == 1            
            
            
def ValidarDatos(PlanillaNombre, PlanillaRC,PlanillaMarca,PlanillaColores,PlanillaColorPrinicipal,PlanillaMemoria,PlanillaDS,PlanillaTecnologia,equipo,equipo_online,loyalty):
    NombreSistema = equipo[1]
    RevenueCodeSis = equipo[4]
    MarcaSis = equipo[11]
    ColoresSis = equipo[13]
    ColorSis=equipo[12]
    MemoriaSis = equipo[8]
    DSSIS=equipo[6]
    TeconolgoiaSis=equipo[7]
    ModeloSis =equipo[5]
    Nrc = equipo[14]
    respuesta = ['OK', '']
    BulkTree = None
    if (loyalty != None):
        CadenaLoyalty = loyalty[5]
    else:
        CadenaLoyalty = None
            
    # print(NombreSistema)
    # print(RevenueCodeSis)
    # print(MarcaSis)
    # print(ColoresSis)
    # print(ColorSis)
    # print(MemoriaSis)
    # print(DSSIS)
    # print(PlanillaDS)
    # print(TeconolgoiaSis)
    # print(ModeloSis)
    # print(Nrc)
    # print(PlanillaNombre)
    
                    
    if(equipo_online): #En caso de ser Online validar el equipment group.
        if (str(equipo[17]) != "None"):
            respuesta[0] = "Error"
            respuesta[1] = "Product Offering Associations erroneo"
        if (not (str(equipo[15]) in NombreSistema)):
            respuesta[0] = "Error"
            if(respuesta[1]):
                respuesta[1] += ", equipment group erroneo -> "
                respuesta[1]+= str(equipo[15])                
            else:
                respuesta[1] = "Equipment group erroneo ->"
                respuesta[1]+= str(equipo[15])         
    else:
        BulkTree = equipo[15]
        if(not validar_clearing(BulkTree)):
            respuesta[0] = "Error"
            if(respuesta[1]):
                respuesta[1] += ", Product Offering Associations erroneo" 
            else:
                respuesta[1] = "Product Offering Associations erroneo"
        if(not (("Empresas Standard:3.33" in str(CadenaLoyalty)) and ("Residencial Standard:3.33" in str(CadenaLoyalty)))):
            respuesta[0] = "Error"
            if(respuesta[1]):
                respuesta[1] += ", Loyalty erroneo deberia ser Empresas Standard y Residencial Standard con One Time Redemption Rate valiendo 3.33"
            else:
                respuesta[1] = "Loyalty erroneo deberia ser Empresas Standard y Residencial Standard con One Time Redemption Rate valiendo 3.33"        
                                                                                
        
    if(str(NombreSistema).lower() != PlanillaNombre):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", Distinto nombre ->"
            respuesta[1]+= str(PlanillaNombre)
        else: 
            respuesta[1] = "Distinto nombre->"
            respuesta[1]+= str(PlanillaNombre)    
    if(str(RevenueCodeSis).lower() != PlanillaRC):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", distinto RC->"
            respuesta[1]+= str(PlanillaRC)
        else:
            respuesta[1] = "Revenue Code distinto->"
            respuesta[1]+= str(PlanillaRC)                
    if(str(ColorSis).lower() != PlanillaColorPrinicipal):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", distinto Default Color->"
            respuesta[1]+= str(PlanillaColorPrinicipal) 
        else:
            respuesta[1] = "Color principal distinto"
            respuesta[1]+= str(PlanillaColorPrinicipal)                                                        
    if set(separarColores(ColoresSis)) != set(separarColores(PlanillaColores)):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", falla colores"
        else:
            respuesta[1] = "Colores definidos distintos"     
    if(str(MarcaSis).lower() != PlanillaMarca):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", distinta Marca->"
            respuesta[1]+= str(PlanillaMarca)
        else:
            respuesta[1] = "Marca distinta->" 
            respuesta[1]+= str(PlanillaMarca)               
    if((PlanillaMemoria == None) or (int(MemoriaSis) != int(PlanillaMemoria))):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", distinta Memoria o Memoria planilla vacia->"
            respuesta[1]+= str(PlanillaMemoria)
        else:
            respuesta[1] = "Memoria distinta o Memoria planilla vacia->"
            respuesta[1]+= str(PlanillaMemoria)                    
    if((DSSIS) != PlanillaDS):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", distinta DS->"
            respuesta[1]+= str(PlanillaDS)
        else:
            respuesta[1] = "DS distinto->"
            respuesta[1]+= str(PlanillaDS)
    if(str(TeconolgoiaSis).lower() != PlanillaTecnologia):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", distinta Tecnologia->"
            respuesta[1]+= str(PlanillaTecnologia)
        else:
            respuesta[1] = "Tecnologia distinto->"
            respuesta[1]+= str(PlanillaTecnologia)
    if str(ModeloSis).lower() != str(PlanillaRC)[-5:].lower():
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", distinto Modelo"
        else:
            respuesta[1] = "Modelo distinto"
    if(not validar_nrc_unico(Nrc)):
        respuesta[0] = "Error"
        if(respuesta[1]):
            respuesta[1] += ", NRC duplicado"
        else:
            respuesta[1] = "NRC duplicado"
    return respuesta    

def imprimirEquipos():
    ''' Esta funcion crear el archivo excel con los equipos'''
    global workbook2, sheet2
    ImprimirEncabezado()
    fila= 2
    salto = 4
    numeroInicio = 2
    Font = "font: name Calibri, color-index black, height 220"
    j=1
    for i in range(len(lista_planilla) - 2):
        i = i + 2
        if(str(lista_planilla[i][buscarPosiciones("Detalle Modelo")]) != "NULL"):
            equipo_nombre = lista_planilla[i][buscarPosiciones("Detalle Modelo")]
            #print(equipo_nombre)           
            revenue_code = (lista_planilla[i][buscarPosiciones("Codigo Equipo RI")])
            #print(revenue_code)
            marca = lista_planilla[i][buscarPosiciones("Marca")]
            #print(marca)
            colores = lista_planilla[i][buscarPosiciones("Colores")]
            #print(colores)
            prin_color = lista_planilla[i][buscarPosiciones("Default Color")]
            #print(prin_color)
            memoria = lista_planilla[i][buscarPosiciones("Memory")]
            #print(memoria)
            dual_sim = lista_planilla[i][buscarPosiciones("Dual SIM")]
            #print(dual_sim)
            tecnologia = lista_planilla[i][buscarPosiciones("Tecnologia")]
            #print(tecnologia)
            for_partners = False
            equipo_online = False
            loyalty = None
            if("online" in str(equipo_nombre).lower()):
                equipo = buscarEquipoOnline(equipo_nombre)
                equipo_online = True
            else: 
                if("partners" in str(equipo_nombre).lower()):
                    equipo = buscarEquipoFP(equipo_nombre)
                    for_partners = True
                else:
                    equipo = buscarEquipo(equipo_nombre)
                    loyalty = buscarLoyalty(equipo_nombre)    
            if(equipo == None):
                print('No se encontro el equipo')
            else:
                ImprimirBase(equipo_nombre, equipo[1], revenue_code, prin_color,colores, j)  
                if(for_partners):
                    #ImprimirBaseForPartners(equipo_nombre,equipo[5],revenue_code,prin_color,colores,j)
                    Resultado = ValidarDatosForPartners(str(equipo_nombre).lower(),str(revenue_code).lower(),str(colores).lower(),str(prin_color).lower(),equipo)
                else:    
                    # ImprimirBase(equipo_nombre, equipo[4], revenue_code, equipo[12],equipo[13], j)
                    Resultado = ValidarDatos(str(equipo_nombre).lower(),str(revenue_code).lower(),str(marca).lower(),colores,str(prin_color).lower(),memoria,dual_sim,str(tecnologia).lower(),equipo,equipo_online,loyalty)
                    
                sheet2.write(j, 5, Resultado[0]) #Status
                sheet2.write(j, 6, Resultado[1]) #Description Error
                j = j+1

    ruta = saveFileDialog()
    workbook2.save(ruta)
    guardarComoXLSX(ruta)


def guardarArchivoText(nombre, objeto):
    '''Funcion guarda archivo en formato texto'''
    archivo = open(nombre,'w') 
    str1 = ',\n'.join(str(e) for e in objeto)
    archivo.write(str1)  
    archivo.close() 
    
    
    
if __name__ == '__main__':
    """Esta la funcion principal"""  
    opcion = 0;
        # Seleccionar opcion para ejecutar script 
    print ('Por favor seleccione alguna de las dos opciones siguientes:')
    print ('1 - Validar caracterisitcas de equipos en Desarrollo')
    print ('2 - Validar caracterisitcas de equipos Activos')
    print ('Cualquier otra tecla para salir')
    print ('Opcion:')
    opcion = input() 
    
    if int(opcion) == 1:
        estado = 'en desarrollo'
        #Consulta de los planes en su primera version en la BD TEST
        lista_id_equipos_TEST = Equipos(estado)
        lista_id_equiposOnline_TEST = Online(estado)
        lista_id_equiposFP_TEST = EquiposFP(estado)
        lista_loyalty = Loyalty(estado)
        now = datetime.datetime.now()
        fecha = now.strftime("%Y-%m-%d %H.%M")
        archivo2 =  'lista de Equipos en TEST '+fecha+'.txt'
        archivo3 = 'Lista de Online en TEST'+fecha+'.txt'
        archivo4 = 'Lista de EquiposFP en TEST'+fecha+'.txt'
        #Guardar lista de planes en archivo
        guardarArchivoText(archivo2, lista_id_equipos_TEST)
        guardarArchivoText(archivo3, lista_id_equiposOnline_TEST)
        guardarArchivoText(archivo4, lista_id_equiposFP_TEST)
        #Solicitar excel de nuevos precios
        ruta = openFileDialog()
        lista_planilla = readCSVPlanes(ruta)
    if int(opcion) == 2:
        estado = 'activa'
        #Consulta de los planes en su primera version en la BD TEST
        lista_id_equipos_TEST = Equipos(estado)
        lista_id_equiposOnline_TEST = Online(estado)
        lista_id_equiposFP_TEST = EquiposFP(estado)
        lista_loyalty = Loyalty(estado)
        now = datetime.datetime.now()
        fecha = now.strftime("%Y-%m-%d %H.%M")
        # archivo2 =  'lista de Equipos en TEST '+fecha+'.txt'
        # archivo3 = 'Lista de Online en TEST'+fecha+'.txt'
        # archivo4 = 'Lista de EquiposFP en TEST'+fecha+'.txt'
        # #Guardar lista de planes en archivo
        # guardarArchivoText(archivo2, lista_id_equipos_TEST)
        # guardarArchivoText(archivo3, lista_id_equiposOnline_TEST)
        # guardarArchivoText(archivo4, lista_id_equiposFP_TEST)
        #Solicitar excel de nuevos precios
        ruta = openFileDialog()
        lista_planilla = readCSVPlanes(ruta)
    if int(opcion) == 1:
        imprimirEquipos()
    if   int(opcion) == 2:
        imprimirEquipos()
    else:
        print ('Saliendo del script...')
    print ('Fin')    
       